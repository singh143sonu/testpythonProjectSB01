from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import pandas as pd
from simple_salesforce import Salesforce

sf = Salesforce(password='*4Gu0kqbL*z8O&XE22', username='sandeep.bhadauria@genesisenergy.co.nz.qa21',
                organizationId='00D0n0000001SP6',domain='test')

#sf = Salesforce(username='sandeep.bhadauria@genesisenergy.co.nz.qa21', password='*4Gu0kqbL*z8O&XE22', security_token='bhind')
print(sf)

root = tk.Tk()

canvas1 = tk.Canvas(root, width=300, height=300, bg='lightsteelblue2', relief='raised')
canvas1.pack()

label1 = tk.Label(root, text='SCV data upload tool', bg='lightsteelblue2')
label1.config(font=('helvetica', 20))
canvas1.create_window(150, 60, window=label1)
read_file_excel = None
read_file_csv = None
import_file_path = None
export_file_path = None
type_of_file_csv = None
type_of_file_excel = None

def getFile():
    global read_file_excel
    global read_file_csv
    global import_file_path
    global type_of_file_excel
    global type_of_file_csv
    import_file_path = filedialog.askopenfilename()
    type_of_file_csv = import_file_path[:-4:-1]         # it should start with vsc.

    if type_of_file_csv == 'vsc':
        read_file_csv = pd.read_csv(import_file_path)
    else:
        MsgBox = tk.messagebox.showinfo('Incorrect file selection', 'Please select a valid .csv file', icon='warning')

#----------------------------
def convertAndUploadFile():
    global export_file_path
    export_file_path = import_file_path[:-4]

    read_file = pd.read_csv(import_file_path)
    read_file.to_excel(export_file_path + '.xlsx', index=None, header=True)

    filename=(export_file_path + ".xlsx")
    wb = load_workbook(filename)
    ws = wb.active

    for i in range(1,ws.max_row+1):
        cell_v = ws.cell(row=i,column=3)
        cell_stripped_value = (cell_v.value.strip())
        cell_v.value = cell_stripped_value

    for src, dst in zip(ws['C:C'],ws['C:C']):
        if src.value == "Small" or src.value == "Medium":
            dst.value = src.value + " Business"
        else:
            dst.value = src.value
    wb.save(filename=filename)

    # data = [
    #     {'Id': '0000000000AAAAA', 'Email': 'examplenew2@example.com'},
    #     {'Email': 'foo@foo.com'}
    # ]
    # sf.bulk.Contact.upsert(data, 'Id', batch_size=10000, use_serial=True)
    for i in range(2, ws.max_row + 1):
        sf.Single_Customer_View.upsert('ACCOUNT_NO__c/'+ A[i], {'Annual_Spend__c': B[i], 'Business_Segment__c': C[i]})


def exitApplication():
    MsgBox = tk.messagebox.askquestion('Exit Application', 'Are you sure you want to exit the application',
                                       icon='warning')
    if MsgBox == 'yes':
        root.destroy()
#-------------------------------
browseButton_Excel = tk.Button(text="  Select CSV File to upload  ", command=getFile, bg='green', fg='white',
                               font=('helvetica', 12, 'bold'))
canvas1.create_window(150, 100, window=browseButton_Excel)
#-------------------------------
saveAsButton_CSV = tk.Button(text='   Upload to Salesforce   ', command=convertAndUploadFile, bg='green', fg='white',
                             font=('helvetica', 12, 'bold'))
canvas1.create_window(150, 150, window=saveAsButton_CSV)
#-------------------------------
exitButton = tk.Button(root, text='       Exit Application     ', command=exitApplication, bg='brown', fg='white',
                       font=('helvetica', 12, 'bold'))
canvas1.create_window(150, 250, window=exitButton)
#-------------------------------

root.mainloop()

